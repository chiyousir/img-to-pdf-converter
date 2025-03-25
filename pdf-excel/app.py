import os
import base64
import uuid
import json
import io
from flask import Flask, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models
import pandas as pd
import tempfile
import openpyxl
from openpyxl import Workbook, load_workbook
import zipfile
from copy import copy
import os.path
import shutil
import hashlib
import platform
import time
import logging
import requests
from openpyxl.utils.cell import get_column_letter
import numpy as np
from PIL import Image
from pdf2image import convert_from_bytes

app = Flask(__name__)

# Configure upload folders
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
TEMP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(TEMP_FOLDER, exist_ok=True)

# Tencent Cloud OCR configuration
SECRET_ID = os.environ.get("TENCENT_SECRET_ID", "your_secret_id_here")
SECRET_KEY = os.environ.get("TENCENT_SECRET_KEY", "your_secret_key_here")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        # Check if file was uploaded
        if 'pdf' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file found in upload'})
        
        file = request.files['pdf']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'})
        
        # Check file type
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'status': 'error', 'message': 'Only PDF files are supported'})
        
        # Check file size (limit to 20MB)
        MAX_SIZE = 20 * 1024 * 1024  # 20MB
        if request.content_length > MAX_SIZE:
            return jsonify({'status': 'error', 'message': 'File too large, maximum size is 20MB'})
        
        # Save uploaded PDF file
        # Ensure filename is not empty and secure
        filename = secure_filename(file.filename) if file.filename else "untitled.pdf"
        file_id = str(uuid.uuid4())
        pdf_path = os.path.join(UPLOAD_FOLDER, f"{file_id}_{filename}")
        file.save(pdf_path)
        
        # Convert PDF to images
        images = convert_from_path(pdf_path)
        
        # If no pages, return error
        if len(images) == 0:
            return jsonify({'status': 'error', 'message': 'No pages detected in PDF file'})
        
        # Generate Excel filename
        base_filename = os.path.splitext(filename)[0]  # Get filename without extension
        excel_filename = f"{base_filename}.xlsx"
        
        # Check if file already exists, add numeric suffix if needed
        counter = 1
        original_excel_filename = excel_filename
        while os.path.exists(os.path.join(OUTPUT_FOLDER, excel_filename)):
            excel_filename = f"{base_filename}_{counter}.xlsx"
            counter += 1
        
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        # Multi-page PDF processing
        # Create a temporary directory to store intermediate files
        temp_dir = tempfile.mkdtemp()
        try:
            # First page PDF processing
            first_image = images[0]
            image_path = os.path.join(temp_dir, "page_1.jpg")
            first_image.save(image_path, 'JPEG')
            
            # Call Tencent Cloud OCR API to recognize tables and get Base64 Excel data for first page
            excel_base64 = recognize_table(image_path)
            
            if not excel_base64:
                return jsonify({'status': 'error', 'message': 'No table data detected in first page'})
            
            # Decode Base64 data from first page to Excel binary data
            excel_binary = base64.b64decode(excel_base64)
            
            # Save first page Excel file (containing Sheet1)
            temp_excel_path = os.path.join(temp_dir, "page_1.xlsx")
            with open(temp_excel_path, 'wb') as f:
                f.write(excel_binary)
            
            # Load first page Excel as base file
            wb = load_workbook(temp_excel_path)
            
            # Rename first worksheet if needed
            if len(wb.sheetnames) > 0:
                # Rename first worksheet to "Page1"
                wb.worksheets[0].title = "Page1"
            
            # Process remaining pages (starting from second page)
            for i in range(1, len(images)):
                page_num = i + 1
                image_path = os.path.join(temp_dir, f"page_{page_num}.jpg")
                images[i].save(image_path, 'JPEG')
                
                # Call Tencent Cloud OCR API to recognize tables and get Base64 Excel data
                excel_base64 = recognize_table(image_path)
                
                if excel_base64:
                    # Decode Base64 data to Excel binary data
                    excel_binary = base64.b64decode(excel_base64)
                    
                    # Save as temporary Excel file
                    temp_excel_path = os.path.join(temp_dir, f"page_{page_num}.xlsx")
                    with open(temp_excel_path, 'wb') as f:
                        f.write(excel_binary)
                    
                    # Load temporary Excel file
                    temp_wb = load_workbook(temp_excel_path)
                    temp_ws = temp_wb.active
                    
                    # Create new worksheet
                    new_sheet_name = f"Page{page_num}"
                    new_ws = wb.create_sheet(title=new_sheet_name)
                    
                    # Collect merged cells information
                    merged_cells = list(temp_ws.merged_cells.ranges)
                    
                    # Copy all cells directly, preserving content and formatting
                    for row in temp_ws.iter_rows():
                        for cell in row:
                            # Copy all cells, including empty ones
                            new_cell = new_ws.cell(row=cell.row, column=cell.column)
                            new_cell.value = cell.value
                            
                            # Copy style attributes
                            if hasattr(cell, 'font'):
                                new_cell.font = copy(cell.font)
                            if hasattr(cell, 'border'):
                                new_cell.border = copy(cell.border)
                            if hasattr(cell, 'fill'):
                                new_cell.fill = copy(cell.fill)
                            if hasattr(cell, 'number_format'):
                                new_cell.number_format = cell.number_format
                            if hasattr(cell, 'protection'):
                                new_cell.protection = copy(cell.protection)
                            if hasattr(cell, 'alignment'):
                                new_cell.alignment = copy(cell.alignment)
                    
                    # Copy merged cells
                    for merged_cell in merged_cells:
                        new_ws.merge_cells(str(merged_cell))
                    
                    # Copy column widths
                    for col_letter, col_dim in temp_ws.column_dimensions.items():
                        if hasattr(col_dim, 'width') and col_dim.width:
                            new_ws.column_dimensions[col_letter].width = col_dim.width
                    
                    # Copy row heights
                    for row_num, row_dim in temp_ws.row_dimensions.items():
                        if hasattr(row_dim, 'height') and row_dim.height:
                            new_ws.row_dimensions[row_num].height = row_dim.height
                    
                    # Close temporary workbook
                    temp_wb.close()
                    
                    print(f"Successfully added worksheet: {new_sheet_name}")
            
            # Save final Excel file
            wb.save(excel_path)
            wb.close()
            
            # Clean up temporary files
            shutil.rmtree(temp_dir)
            
            # Return download link
            download_url = f"/download/{excel_filename}"
            return jsonify({'status': 'success', 'download_url': download_url})
            
        except Exception as e:
            print(f"Error processing multi-page PDF: {str(e)}")
            # Clean up temporary directory
            shutil.rmtree(temp_dir)
            return jsonify({'status': 'error', 'message': f'Error processing multi-page PDF: {str(e)}'})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error processing file: {str(e)}'})

def merge_excel_files(file_infos, output_path):
    """
    Merge multiple Excel files into one Excel file, each file as a separate worksheet
    
    Args:
        file_infos: List of dictionaries containing path and target worksheet name
        output_path: Output Excel file path
    """
    try:
        # Create a base Excel file (using the first Excel file)
        if not file_infos:
            # If no files, create an empty Excel
            wb = Workbook()
            wb.save(output_path)
            return
            
        # Copy first Excel file as base
        shutil.copy(file_infos[0]['path'], output_path)
        
        # Use zipfile to directly merge other Excel files' worksheets into the base file
        # Excel is actually a zip file containing multiple xml files
        with zipfile.ZipFile(output_path, 'a') as zip_out:
            # Get worksheet info file
            wb = load_workbook(output_path)
            
            # Process each additional Excel file (starting from the second)
            for i, file_info in enumerate(file_infos[1:], start=2):
                # Worksheet name
                sheet_name = file_info['sheet_name']
                
                # Temporary file path
                temp_file = file_info['path']
                
                # Directly add worksheet to existing Excel (via openpyxl)
                # This preserves original formatting and data
                src_wb = load_workbook(temp_file)
                src_sheet = src_wb.active
                
                # Create new worksheet
                if sheet_name in wb.sheetnames:
                    sheet_name = f"{sheet_name}_{i}"  # Avoid duplicate names
                
                # Add worksheet to workbook
                target_sheet = wb.create_sheet(title=sheet_name)
                
                # Copy cell contents and format
                for row in src_sheet.rows:
                    for cell in row:
                        target_sheet[cell.coordinate] = cell.value
                        
                # Copy column widths
                for col_letter, col_dim in src_sheet.column_dimensions.items():
                    if hasattr(col_dim, 'width') and col_dim.width:
                        target_sheet.column_dimensions[col_letter].width = col_dim.width
            
            # Save merged workbook
            wb.save(output_path)
            
    except Exception as e:
        print(f"Error merging Excel files: {str(e)}")
        # If error occurs, at least keep the first Excel file
        if file_infos:
            shutil.copy(file_infos[0]['path'], output_path)

def recognize_table(image_path, max_retries=2):
    """Call Tencent Cloud OCR API to recognize tables and return Base64 encoded Excel data
    
    Args:
        image_path: Image path
        max_retries: Maximum number of retries
        
    Returns:
        Base64 encoded Excel data or None
    """
    retry_count = 0
    last_error = None
    
    while retry_count <= max_retries:
        try:
            # Read image file and convert to Base64
            with open(image_path, 'rb') as f:
                image_data = f.read()
                image_base64 = base64.b64encode(image_data).decode('utf-8')
            
            # Configure Tencent Cloud OCR client
            cred = credential.Credential(SECRET_ID, SECRET_KEY)
            http_profile = HttpProfile()
            http_profile.endpoint = "ocr.tencentcloudapi.com"
            client_profile = ClientProfile()
            client_profile.httpProfile = http_profile
            client = ocr_client.OcrClient(cred, "ap-guangzhou", client_profile)
            
            # 创建表格识别请求
            req = models.RecognizeTableOCRRequest()
            req.ImageBase64 = image_base64
            
            # 发送请求并获取响应
            resp = client.RecognizeTableOCR(req)
            result = json.loads(resp.to_json_string())
            
            # 直接返回API返回的Base64编码Excel数据
            if 'Data' in result and result['Data']:
                return result['Data']
            
            # 如果没有识别到表格数据
            print(f"API调用成功但未返回表格数据: {result}")
            return None
            
        except TencentCloudSDKException as err:
            last_error = err
            print(f"腾讯云OCR API错误 (尝试 {retry_count+1}/{max_retries+1}): {err}")
            retry_count += 1
            if retry_count <= max_retries:
                # 失败后等待一段时间再重试
                time.sleep(1)  # 等待1秒
            else:
                print(f"达到最大重试次数，放弃识别")
                break
        except Exception as e:
            print(f"表格识别错误: {e}")
            last_error = e
            break
    
    print(f"表格识别失败: {last_error}")
    return None

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)